"""Excel and CSV exporters for ClassStartPlan."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook

from startplanner.domain import Competition
from startplanner.services.course_grid import build_course_grid


def _class_sort_maps(
    competition: Competition,
) -> tuple[dict[str, int], dict[str, str]]:
    orders = {rc.id: rc.sort_order for rc in competition.classes.values()}
    names = {rc.id: rc.name for rc in competition.classes.values()}
    return orders, names


def _plan_rows(
    competition: Competition, start_location_id: str | None = None
) -> list[list[str]]:
    header = [
        "Lähtö",
        "Järjestys",
        "1. lähtöaika",
        "Sarja",
        "Kilpailijoita",
        "Lähtöväli",
        "Rata",
        "1. rasti",
    ]
    rows = [header]
    orders, names = _class_sort_maps(competition)
    location_ids = (
        [start_location_id] if start_location_id else sorted(competition.plans.keys())
    )
    for loc_id in location_ids:
        plan = competition.plan_for(loc_id)
        if not plan:
            continue
        loc = competition.start_locations.get(loc_id)
        loc_name = loc.name if loc else loc_id
        for entry in plan.sorted_entries(
            by="class", class_sort_order=orders, class_names=names
        ):
            rc = competition.classes.get(entry.class_id)
            course = competition.course_for_class(rc) if rc else None
            rows.append(
                [
                    loc_name,
                    str(rc.sort_order if rc else ""),
                    entry.first_start_time.strftime("%H:%M"),
                    rc.name if rc else "",
                    str(competition.competitor_count(entry.class_id)),
                    str(rc.start_interval_min if rc else ""),
                    course.name if course else "",
                    (course.first_control if course else "") or "",
                ]
            )
    return rows


def _sheet_title(prefix: str, name: str) -> str:
    raw = f"{prefix} – {name}" if name else prefix
    return raw[:31]


def _write_grid_sheet(wb: Workbook, competition: Competition, loc_id: str) -> None:
    plan = competition.plan_for(loc_id)
    grid = build_course_grid(competition, plan)
    loc = competition.start_locations.get(loc_id)
    loc_name = loc.name if loc else loc_id
    ws = wb.create_sheet(_sheet_title("Ruudukko", loc_name))
    header = ["Aika", "Yht"] + [
        f"{col.course_name} (1. {col.first_control or '—'})" for col in grid.columns
    ]
    ws.append(header)
    for minute in grid.minutes:
        row = [minute.strftime("%H:%M"), grid.total(minute)]
        for col in grid.columns:
            row.append(grid.cell(minute, col.course_id))
        ws.append(row)


class ExcelExporter:
    def write(
        self,
        competition: Competition,
        path: str | Path,
        start_location_id: str | None = None,
    ) -> None:
        from startplanner.services.quality_service import QualityService

        wb = Workbook()
        ws = wb.active
        ws.title = "Lähtökaavio"
        for r_idx, row in enumerate(_plan_rows(competition, start_location_id), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        summary = wb.create_sheet("Yhteenveto")
        summary.append(["Kilpailu", competition.name])
        summary.append(
            [
                "Kilpailun alku",
                competition.settings.competition_start.strftime("%H:%M"),
            ]
        )
        summary.append(
            ["Oletusväli (min)", competition.settings.default_start_interval_min]
        )
        summary.append(["Sarjaväli (min)", competition.settings.class_gap_min])
        summary.append([])
        summary.append(
            ["Lähtö", "Laatu", "Säännöt", "1. rastit", "Virtaus", "Järjestys", "Välit"]
        )
        quality = QualityService()
        loc_ids = (
            [start_location_id]
            if start_location_id
            else sorted(competition.start_locations.keys())
        )
        for loc_id in loc_ids:
            if not competition.plan_for(loc_id):
                continue
            loc = competition.start_locations.get(loc_id)
            score = quality.score(competition, loc_id)
            summary.append(
                [
                    loc.name if loc else loc_id,
                    score.total,
                    score.rules,
                    score.first_controls,
                    score.flow,
                    score.order,
                    score.gaps,
                ]
            )

        grid_locs = (
            [start_location_id]
            if start_location_id
            else sorted(competition.plans.keys())
        )
        for loc_id in grid_locs:
            if competition.plan_for(loc_id):
                _write_grid_sheet(wb, competition, loc_id)

        wb.save(path)


class CsvExporter:
    def write(
        self,
        competition: Competition,
        path: str | Path,
        start_location_id: str | None = None,
    ) -> None:
        with Path(path).open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerows(_plan_rows(competition, start_location_id))
